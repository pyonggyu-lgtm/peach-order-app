# =============================================================================
# 🍑 실크로드농원 주문관리 시스템 (Streamlit + Google Sheets)
# =============================================================================
#
# ┌─────────────────────────────────────────────────────────────────────────┐
# │  【 Google Sheets 최초 세팅 가이드 】                                      │
# │                                                                         │
# │  1단계: Google Cloud 프로젝트 & 서비스 계정 생성                            │
# │  ─────────────────────────────────────────────────────────────────────  │
# │  1. https://console.cloud.google.com 접속 → 새 프로젝트 생성              │
# │  2. 왼쪽 메뉴 "API 및 서비스" → "라이브러리" 클릭                          │
# │  3. "Google Sheets API" 검색 → 활성화                                    │
# │  4. "Google Drive API" 검색 → 활성화                                     │
# │  5. 왼쪽 메뉴 "API 및 서비스" → "사용자 인증 정보" 클릭                    │
# │  6. "사용자 인증 정보 만들기" → "서비스 계정" 선택                          │
# │  7. 서비스 계정 이름 입력 (예: peach-farm) → 완료                          │
# │  8. 생성된 서비스 계정 클릭 → "키" 탭 → "키 추가" → "새 키 만들기"         │
# │  9. JSON 형식 선택 → 다운로드 (이 파일이 서비스 계정 키입니다)              │
# │                                                                         │
# │  2단계: Google Sheets 스프레드시트 생성                                    │
# │  ─────────────────────────────────────────────────────────────────────  │
# │  1. https://sheets.google.com 에서 새 스프레드시트 생성                   │
# │  2. 시트 4개 생성: 주문목록 / 상품목록 / 고객목록 / 설정                   │
# │  3. 주문목록 1행 헤더:                                                    │
# │     주문번호|주문일시|주문자이름|주문자전화번호|주문자이메일|               │
# │     받는분이름|받는분전화번호|받는분주소|상품명|수량|배송메모|상태          │
# │  4. 상품목록 1행: 상품명 | 단가 | 설명                                    │
# │     2행부터 상품 입력 (예: 복숭아 4kg 일반용 | 30000 | 신선한 복숭아)      │
# │  5. 고객목록 1행: 이름 | 이메일 | 전화번호                                 │
# │  6. 설정 시트 A/B열:                                                     │
# │     order_start    | 2026-07-01 09:00                                   │
# │     order_end      | 2026-07-10 18:00                                   │
# │     bank           | 농협                                                │
# │     account_number | 000-0000-0000                                      │
# │     holder         | 장명숙                                              │
# │                                                                         │
# │  3단계: 서비스 계정 이메일로 스프레드시트 공유                              │
# │  ─────────────────────────────────────────────────────────────────────  │
# │  스프레드시트 → 공유 → 서비스 계정 client_email 입력 → 편집자 권한 부여     │
# │                                                                         │
# │  4단계: .streamlit/secrets.toml 작성                                     │
# │  ─────────────────────────────────────────────────────────────────────  │
# │  [gcp_service_account]                                                  │
# │  type = "service_account"                                               │
# │  project_id = "your-project-id"                                         │
# │  private_key_id = "..."                                                 │
# │  private_key = "-----BEGIN RSA PRIVATE KEY-----\n...\n-----END..."      │
# │  client_email = "peach@project.iam.gserviceaccount.com"                │
# │  client_id = "..."                                                      │
# │  auth_uri = "https://accounts.google.com/o/oauth2/auth"                │
# │  token_uri = "https://oauth2.googleapis.com/token"                     │
# │                                                                         │
# │  [app]                                                                  │
# │  admin_password = "복숭아1234"                                           │
# │  spreadsheet_id = "1ABC...xyz"                                          │
# │  farm_name = "실크로드농원"                                          │
# │  order_url = "https://your-app.streamlit.app"                           │
# │                                                                         │
# │  [email]                                                                │
# │  sender = "your-gmail@gmail.com"                                        │
# │  app_password = "xxxx xxxx xxxx xxxx"                                   │
# │                                                                         │
# │  [account]                                                              │
# │  bank = "농협"                                                           │
# │  account_number = "000-0000-0000"                                       │
# │  holder = "장명숙"                                                       │
# │                                                                         │
# │  실행: streamlit run peach_order_system.py                              │
# └─────────────────────────────────────────────────────────────────────────┘
# =============================================================================

import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import io
import random
import string
import re

# CoolSMS — 설치 여부에 따라 동적 import
try:
    from coolsms import CoolSMS as _CoolSMS
    _COOLSMS_AVAILABLE = True
except ImportError:
    _COOLSMS_AVAILABLE = False

# =============================================================================
# 페이지 기본 설정
# =============================================================================
st.set_page_config(
    page_title="복숭아농장 주문관리",
    page_icon="🍑",
    layout="wide",
    initial_sidebar_state="auto",
)

# 스마트폰 핀치줌(두 손가락 확대) 허용
# st.markdown의 <script>는 Streamlit이 보안상 차단하므로 components.html 사용
import streamlit.components.v1 as _components
_components.html("""
<script>
(function() {
    var content = 'width=device-width, initial-scale=1.0, maximum-scale=5.0, user-scalable=yes';
    function fix() {
        try {
            // components.html은 iframe 안에서 실행 → parent.document로 실제 페이지 접근
            var meta = parent.document.querySelector('meta[name="viewport"]');
            if (meta) {
                meta.setAttribute('content', content);
            } else {
                var m = parent.document.createElement('meta');
                m.name = 'viewport'; m.content = content;
                parent.document.head.appendChild(m);
            }
        } catch(e) {}
    }
    fix();
    setTimeout(fix, 300);
    setTimeout(fix, 1000);
})();
</script>
""", height=0)

# =============================================================================
# 전체 테마 CSS — 따뜻한 복숭아 색상
# =============================================================================
st.markdown("""
<style>
/* ── 전체 배경 ── */
[data-testid="stAppViewContainer"] {
    background: linear-gradient(135deg, #fff8f0 0%, #fff3e8 50%, #ffeedd 100%);
}

/* ── 사이드바 ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #ff8c42 0%, #e55a00 100%);
}
[data-testid="stSidebar"] * { color: white !important; }
[data-testid="stSidebar"] .stTextInput input {
    background: rgba(255,255,255,0.95) !important;
    color: #333 !important;
    border: 1px solid rgba(255,255,255,0.6) !important;
    border-radius: 8px !important;
}
[data-testid="stSidebar"] .stTextInput input::placeholder {
    color: #aaa !important;
}

/* ── 헤더 ── */
.peach-header {
    background: linear-gradient(135deg, #ff8c42, #ff6b1a, #e55a00);
    color: white;
    padding: 2rem 1.5rem;
    border-radius: 16px;
    text-align: center;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 20px rgba(255,107,26,0.3);
}
.peach-header h1 {
    font-size: clamp(1.25rem, 5vw, 2.2rem);
    margin: 0;
    line-height: 1.35;
    word-break: keep-all;
    text-shadow: 1px 1px 3px rgba(0,0,0,0.2);
}
.peach-header .admin-sub {
    display: block;
    font-size: 0.65em;
    font-weight: 600;
    opacity: 0.88;
    letter-spacing: 2px;
    margin-top: 2px;
}
.peach-header p  { margin: 0.5rem 0 0 0; opacity: 0.9; font-size: clamp(0.85rem, 3vw, 1.05rem); word-break: keep-all; }

/* ── 카드 ── */
.peach-card {
    background: white;
    border-radius: 12px;
    padding: 1.4rem;
    box-shadow: 0 2px 12px rgba(255,107,26,0.12);
    border-left: 4px solid #ff8c42;
    margin-bottom: 1rem;
}

/* ── 공지 박스 ── */
.notice-box {
    background: #fff3e0;
    border: 2px solid #ffb74d;
    border-radius: 12px;
    padding: 1.4rem 1.5rem;
    margin: 1rem 0;
    text-align: center;
}
.notice-box.closed { background: #fce4ec; border-color: #e57373; }
.notice-box.open   { background: #e8f5e9; border-color: #66bb6a; }

/* ── 계좌 안내 ── */
.account-box {
    background: linear-gradient(135deg, #fff8f0, #fff3e8);
    border: 2px solid #ff8c42;
    border-radius: 12px;
    padding: 1.5rem;
    text-align: center;
    margin: 1rem 0;
}
.account-box .bank-name  { font-size: 1.3rem; font-weight: bold; color: #e55a00; }
.account-box .account-num { font-size: clamp(1.1rem, 5.5vw, 1.8rem); font-weight: bold; color: #333; letter-spacing: 1px; margin: 0.3rem 0; white-space: nowrap; }

/* ── 수령자 박스 ── */
.recipient-box {
    background: #fff8f0;
    border: 1px solid #ffcc99;
    border-radius: 10px;
    padding: 1rem;
    margin-bottom: 0.8rem;
}
.recipient-box-title { font-weight: bold; color: #e55a00; font-size: 1.05rem; margin-bottom: 0.5rem; }

/* ── 카운트다운 ── */
.countdown-box {
    background: linear-gradient(135deg, #e8f5e9, #c8e6c9);
    border: 2px solid #66bb6a;
    border-radius: 12px;
    padding: 1rem;
    text-align: center;
    font-size: clamp(0.9rem, 3.5vw, 1.1rem);
    color: #2e7d32;
    font-weight: bold;
    margin: 0.5rem 0;
    word-break: keep-all;
    line-height: 1.5;
}
/* ── 섹션 제목 (h3) 모바일 자연스러운 줄바꿈 ── */
h3 {
    word-break: keep-all;
    line-height: 1.4;
    font-size: clamp(1rem, 4vw, 1.4rem) !important;
}

/* ── 버튼 ── */
.stButton > button {
    background: linear-gradient(135deg, #ff8c42, #ff6b1a) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.5rem 1.5rem !important;
    font-weight: bold !important;
    transition: all 0.2s ease !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #ff6b1a, #e55a00) !important;
    box-shadow: 0 3px 10px rgba(255,107,26,0.4) !important;
    transform: translateY(-1px) !important;
}

/* ── 탭 ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    background: white;
    border-radius: 10px;
    padding: 4px;
}
.stTabs [data-baseweb="tab"]    { border-radius: 8px; padding: 0.4rem 1rem; }
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #ff8c42, #ff6b1a) !important;
    color: white !important;
}

/* ── 지표 카드 ── */
.metric-card {
    background: white;
    border-radius: 10px;
    padding: 1rem;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    border-top: 3px solid #ff8c42;
}
.metric-card .metric-value { font-size: clamp(0.9rem, 1.7vw, 2rem); font-weight: bold; white-space: nowrap; }
.metric-card .metric-label { color: #666; font-size: 0.9rem; }

/* ── 모바일 반응형 ── */
@media (max-width: 768px) {
    .peach-header { padding: 1.3rem 1rem; }
}

/* ── 모바일 스크롤 / 핀치줌 허용 ── */
* {
    touch-action: pan-y pinch-zoom !important;
}
html, body {
    overflow-y: auto !important;
    -webkit-overflow-scrolling: touch !important;
}
</style>
""", unsafe_allow_html=True)


# =============================================================================
# Google Sheets 연결 헬퍼
# =============================================================================

@st.cache_resource(ttl=300)
def get_gspread_client():
    """
    Google Sheets API 클라이언트를 생성합니다.
    secrets.toml의 [gcp_service_account] 항목을 읽습니다.
    연결 실패 시 None을 반환합니다 (앱이 중단되지 않음).
    """
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=scopes,
        )
        return gspread.authorize(creds)
    except Exception:
        return None


def get_spreadsheet():
    """스프레드시트 객체를 반환합니다. 실패 시 None."""
    client = get_gspread_client()
    if client is None:
        return None
    try:
        return client.open_by_key(st.secrets["app"]["spreadsheet_id"])
    except Exception:
        return None


def get_sheet(name: str):
    """시트 이름으로 워크시트 객체를 반환합니다. 실패 시 None."""
    ss = get_spreadsheet()
    if ss is None:
        return None
    try:
        return ss.worksheet(name)
    except Exception:
        return None


# =============================================================================
# 설정 시트 읽기 / 쓰기
# =============================================================================

@st.cache_data(ttl=60)
def load_settings() -> dict:
    """
    '설정' 시트에서 key|value 형태로 설정값을 읽습니다.
    시트 연결 실패 시 secrets.toml의 [account] 값을 기본값으로 사용합니다.
    """
    # secrets에서 기본값 읽기 (secrets.toml 없어도 기본값으로 동작)
    try:
        def_bank  = st.secrets["account"]["bank"]
        def_acct  = st.secrets["account"]["account_number"]
        def_holder = st.secrets["account"]["holder"]
    except Exception:
        def_bank, def_acct, def_holder = "농협", "000-0000-0000", "장명숙"

    defaults = {
        "order_start":    "2026-07-01 09:00",
        "order_end":      "2026-07-10 18:00",
        "bank":           def_bank,
        "account_number": def_acct,
        "holder":         def_holder,
        "farm_phone":     "",
        "farm_address":   "",
    }

    sheet = get_sheet("설정")
    if sheet is None:
        return defaults

    try:
        rows = sheet.get_all_values()
        settings = defaults.copy()
        for row in rows:
            if len(row) >= 2 and row[0].strip():
                settings[row[0].strip()] = row[1].strip()
        # 은행명 정규화: "농협은행", "NH농협은행", "농협 은행" 등 → "농협"
        # (표시 시 자동으로 "은행"이 덧붙으므로 끝의 "은행" 표기를 제거해 중복 방지)
        bank_val = settings.get("bank", "").strip()
        if "농협" in bank_val:
            bank_val = "농협"
        elif bank_val.endswith("은행"):
            bank_val = bank_val[:-2].strip()
        settings["bank"] = bank_val
        return settings
    except Exception:
        return defaults


def save_settings(settings: dict) -> bool:
    """설정 딕셔너리를 '설정' 시트에 key|value 형태로 저장합니다."""
    sheet = get_sheet("설정")
    if sheet is None:
        return False
    try:
        sheet.clear()
        rows = [[k, v] for k, v in settings.items()]
        if rows:
            sheet.update("A1", rows)
        return True
    except Exception as e:
        st.error(f"설정 저장 실패: {e}")
        return False


# =============================================================================
# 상품 목록
# =============================================================================

@st.cache_data(ttl=60)
def load_products() -> list:
    """'상품목록' 시트 A열(2행~)에서 상품명을 읽어 반환합니다."""
    sheet = get_sheet("상품목록")
    if sheet is None:
        # 시트 연결 전 기본 상품 (데모용)
        return ["복숭아 4kg 선물용", "복숭아 4kg 일반용"]
    try:
        rows = sheet.get_all_values()
        # 1행은 헤더(상품명|단가|설명), 2행부터 데이터
        products = [row[0].strip() for row in rows[1:] if row and row[0].strip()]
        if not products:
            return ["복숭아 4kg 선물용", "복숭아 4kg 일반용"]
        # 선물용이 일반용보다 위에 오도록 정렬
        products.sort(key=lambda n: (0 if "선물용" in n else 1 if "일반용" in n else 2))
        return products
    except Exception:
        return ["복숭아 4kg 선물용", "복숭아 4kg 일반용"]


@st.cache_data(ttl=60)
def load_product_prices() -> dict:
    """'상품목록' 시트 B열(단가)에서 상품명→단가 딕셔너리를 반환합니다."""
    sheet = get_sheet("상품목록")
    if sheet is None:
        return {}
    try:
        rows = sheet.get_all_values()
        prices = {}
        for row in rows[1:]:
            if row and row[0].strip():
                name = row[0].strip()
                price = 0
                if len(row) > 1 and row[1].strip():
                    try:
                        price = int(str(row[1]).replace(",", "").replace("원", "").strip())
                    except ValueError:
                        price = 0
                prices[name] = price
        return prices
    except Exception:
        return {}


@st.cache_data(ttl=60)
def load_products_full() -> pd.DataFrame:
    """'상품목록' 시트 전체를 [상품명, 단가, 설명, 상태] DataFrame으로 반환합니다."""
    cols = ["상품명", "단가", "설명", "상태"]
    sheet = get_sheet("상품목록")
    if sheet is None:
        return pd.DataFrame(columns=cols)
    try:
        rows = sheet.get_all_values()
        if len(rows) < 2:
            return pd.DataFrame(columns=cols)
        df = pd.DataFrame(rows[1:], columns=rows[0]).fillna("")
        df = df[[c for c in df.columns if str(c).strip()]]
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        df = df[cols].astype(str)
        # 빈 값(None/nan 문자열 포함)을 빈칸으로 정리하고, 상품명 없는 줄은 제외
        df = df.replace({"None": "", "nan": "", "NaN": ""})
        # 상태가 비어 있으면 '판매중'으로 기본 표시
        df.loc[df["상태"].str.strip() == "", "상태"] = "판매중"
        df = df[df["상품명"].str.strip() != ""].reset_index(drop=True)
        return df
    except Exception:
        return pd.DataFrame(columns=cols)


@st.cache_data(ttl=60)
def load_products_struct() -> list:
    """
    상품을 시트 순서대로 구조화해 반환합니다.
    각 항목: {상품명, 단가, 상태, 품종, 용도}
    - 품종 = 상품명 첫 단어(예: '아마쯔쿠시 4kg 선물용' → '아마쯔쿠시')
    - 용도 = 선물용 / 일반용
    - 상태 = 판매중 / 마감 (없으면 판매중)
    """
    def _mk(name, price=0, status="판매중"):
        name = str(name).strip()
        toks = name.split()
        variety = toks[0] if toks else name
        usage = "선물용" if "선물용" in name else ("일반용" if "일반용" in name else "기타")
        return {"상품명": name, "단가": price, "상태": (status or "판매중").strip() or "판매중",
                "품종": variety, "용도": usage}

    default = [_mk("수황 4kg 선물용", 40000), _mk("수황 4kg 일반용", 35000)]
    sheet = get_sheet("상품목록")
    if sheet is None:
        return default
    try:
        rows = sheet.get_all_values()
        if len(rows) < 2:
            return default
        header = [str(h).strip() for h in rows[0]]
        ci_name   = header.index("상품명") if "상품명" in header else 0
        ci_price  = header.index("단가")   if "단가"   in header else 1
        ci_status = header.index("상태")   if "상태"   in header else None
        items = []
        for row in rows[1:]:
            if len(row) <= ci_name:
                continue
            name = str(row[ci_name]).strip()
            if not name:
                continue
            price = 0
            if ci_price is not None and len(row) > ci_price and str(row[ci_price]).strip():
                try:
                    price = int(str(row[ci_price]).replace(",", "").replace("원", "").strip())
                except ValueError:
                    price = 0
            status = "판매중"
            if ci_status is not None and len(row) > ci_status and str(row[ci_status]).strip():
                status = str(row[ci_status]).strip()
            items.append(_mk(name, price, status))
        return items or default
    except Exception:
        return default


def _render_product_qtys(products_struct, key_prefix):
    """
    품종별로 묶어 수량 입력칸을 렌더링합니다.
    - 품종 순서는 상품목록 시트의 등장 순서 유지
    - 품종 안에서 선물용 → 일반용 순
    - 상태가 '마감'이면 회색(품절)으로 입력 차단
    위젯 key = f"{key_prefix}{상품명}", 반환값 = {상품명: 수량}
    """
    qtys = {}
    varieties = []
    for p in products_struct:
        if p["품종"] not in varieties:
            varieties.append(p["품종"])
    for v in varieties:
        group = [p for p in products_struct if p["품종"] == v]
        group.sort(key=lambda p: 0 if p["용도"] == "선물용" else 1 if p["용도"] == "일반용" else 2)
        all_closed  = all(p["상태"] == "마감" for p in group)
        all_unavail = all(p["상태"] in ("마감", "예정") for p in group)
        if all_closed:
            head_suffix = "　—　마감(품절)"
        elif all_unavail:
            head_suffix = "　—　판매예정"
        else:
            head_suffix = ""
        head = f"🍑 {v}" + head_suffix
        st.markdown(f"**{head}**")
        for p in group:
            name = p["상품명"]
            toks = name.split(maxsplit=1)
            short = toks[1] if len(toks) > 1 else name   # 예: '4kg 선물용'
            if p["상태"] == "마감":
                st.number_input(f"{short}  (마감)", min_value=0, max_value=0,
                                value=0, step=1, key=f"{key_prefix}{name}", disabled=True)
                qtys[name] = 0
            elif p["상태"] == "예정":
                st.number_input(f"{short}  (판매예정)", min_value=0, max_value=0,
                                value=0, step=1, key=f"{key_prefix}{name}", disabled=True)
                qtys[name] = 0
            else:
                qtys[name] = st.number_input(f"{short}  (박스)", min_value=0, max_value=99,
                                             value=0, step=1, key=f"{key_prefix}{name}")
    return qtys


def save_products(df: pd.DataFrame) -> bool:
    """편집된 상품 DataFrame을 '상품목록' 시트에 저장합니다 (헤더 포함 전체 교체)."""
    sheet = get_sheet("상품목록")
    if sheet is None:
        return False
    try:
        rows = [["상품명", "단가", "설명", "상태"]]
        for _, r in df.iterrows():
            name = str(r.get("상품명", "") or "").strip()
            if not name:
                continue   # 상품명 없는 빈 줄은 건너뜀
            raw = str(r.get("단가", "") or "").replace(",", "").replace("원", "").strip()
            try:
                price = int(float(raw)) if raw else 0
            except ValueError:
                price = 0
            desc   = str(r.get("설명", "") or "").strip()
            status = str(r.get("상태", "") or "").strip() or "판매중"
            rows.append([name, price, desc, status])
        sheet.clear()
        sheet.update("A1", rows)
        return True
    except Exception as e:
        st.error(f"상품 저장 실패: {e}")
        return False


# =============================================================================
# 주문 저장 / 로드
# =============================================================================

def generate_order_number() -> str:
    """주문번호 생성: PEACH-YYYYMMDD-XXXXXX (6자리 영숫자 대문자)"""
    today  = datetime.now().strftime("%Y%m%d")
    suffix = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"PEACH-{today}-{suffix}"


def save_orders(order_rows: list) -> bool:
    """
    '주문목록' 시트에 주문 행을 추가합니다.
    시트가 비어있으면 헤더를 먼저 작성합니다.
    """
    sheet = get_sheet("주문목록")
    if sheet is None:
        return False
    try:
        existing = sheet.get_all_values()
        if not existing:
            # 헤더 자동 생성 (12컬럼 — _submit_order와 동일한 스키마)
            header = [
                "주문번호", "주문일시", "주문자이름", "주문자전화번호", "주문자주소",
                "받는분이름", "받는분전화번호", "받는분주소",
                "상품명", "수량", "배송메모", "상태",
            ]
            sheet.append_row(header)
        # 한 번의 API 호출로 전체 행을 일괄 저장
        sheet.append_rows(order_rows)
        return True
    except Exception as e:
        st.error(f"주문 저장 중 오류가 발생했습니다: {e}")
        return False


@st.cache_data(ttl=30)
def load_orders() -> pd.DataFrame:
    """'주문목록' 시트 전체를 DataFrame으로 반환합니다."""
    sheet = get_sheet("주문목록")
    if sheet is None:
        return pd.DataFrame()
    try:
        rows = sheet.get_all_values()
        if len(rows) < 2:
            return pd.DataFrame()
        df = pd.DataFrame(rows[1:], columns=rows[0])
        # Google Sheets 사용 범위가 실제 데이터 범위를 초과할 경우
        # 빈 이름 컬럼(Arrow 직렬화 오류 원인)과 빈 행을 제거
        df = df[[c for c in df.columns if str(c).strip()]]
        # 중복 컬럼명 제거 (첫 번째만 유지) — 중복 시 df["상태"] 가 DataFrame 반환되어 TypeError 발생
        df = df.loc[:, ~df.columns.duplicated(keep="first")]
        if "주문번호" in df.columns:
            df = df[df["주문번호"].str.strip() != ""].reset_index(drop=True)
        return df
    except Exception:
        return pd.DataFrame()


# =============================================================================
# 고객 목록
# =============================================================================

def load_customers() -> pd.DataFrame:
    """'고객목록' 시트에서 고객 정보를 DataFrame으로 반환합니다."""
    sheet = get_sheet("고객목록")
    if sheet is None:
        return pd.DataFrame(columns=["이름", "이메일", "전화번호"])
    try:
        rows = sheet.get_all_values()
        if len(rows) < 2:
            return pd.DataFrame(columns=["이름", "이메일", "전화번호"])
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df = df[[c for c in df.columns if str(c).strip()]]
        return df
    except Exception:
        return pd.DataFrame(columns=["이름", "이메일", "전화번호"])


# =============================================================================
# 이메일 발송
# =============================================================================

def send_email(to_addr: str, subject: str, body: str) -> bool:
    """
    Gmail SMTP(SSL 465포트)를 통해 이메일을 발송합니다.
    secrets.toml의 [email] sender / app_password 를 사용합니다.
    Gmail 앱 비밀번호는 Google 계정 → 보안 → 2단계 인증 → 앱 비밀번호에서 생성합니다.
    """
    try:
        sender   = st.secrets["email"]["sender"]
        app_pw   = st.secrets["email"]["app_password"]
    except Exception:
        # secrets 미설정 시 발송 스킵 (주문은 이미 저장됨)
        return False

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = sender
        msg["To"]      = to_addr
        html_body = "<html><body>" + body.replace("\n", "<br>") + "</body></html>"
        msg.attach(MIMEText(body,      "plain", "utf-8"))
        msg.attach(MIMEText(html_body, "html",  "utf-8"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, app_pw)
            server.sendmail(sender, to_addr, msg.as_string())
        return True
    except Exception as e:
        st.warning(f"이메일 발송 실패 ({to_addr}): {e}")
        return False


# =============================================================================
# SMS 발송 (CoolSMS)
# =============================================================================

def send_sms(to_number: str, message: str) -> bool:
    """
    CoolSMS를 통해 SMS를 발송합니다.
    secrets.toml의 [coolsms] api_key / api_secret / sender 를 사용합니다.

    secrets.toml 설정 예시:
        [coolsms]
        api_key    = "NCSXXXXXXXXXXXXXX"
        api_secret = "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        sender     = "01012345678"   # 발신번호 (CoolSMS에 등록된 번호)
    """
    if not _COOLSMS_AVAILABLE:
        return False          # 패키지 미설치 시 조용히 스킵
    try:
        api_key    = st.secrets["coolsms"]["api_key"]
        api_secret = st.secrets["coolsms"]["api_secret"]
        sender     = st.secrets["coolsms"]["sender"]
    except Exception:
        return False          # secrets 미설정 시 스킵

    try:
        sms = _CoolSMS(api_key, api_secret)
        sms.send({
            "to":   re.sub(r"\D", "", to_number),   # 숫자만 전달
            "from": re.sub(r"\D", "", sender),
            "text": message,
            "type": "SMS" if len(message) <= 90 else "LMS",
        })
        return True
    except Exception as e:
        st.warning(f"문자 발송 실패 ({to_number}): {e}")
        return False


def send_sms_bulk(numbers: list, message: str) -> tuple:
    """
    여러 번호에 동일 메시지를 일괄 발송합니다.
    반환: (성공 건수, 실패 건수)
    """
    ok, fail = 0, 0
    for num in numbers:
        if send_sms(num, message):
            ok += 1
        else:
            fail += 1
    return ok, fail


# =============================================================================
# 주문 기간 판단
# =============================================================================

def check_order_period(settings: dict):
    """
    현재 시각과 주문 기간을 비교합니다.
    반환: ("before" | "open" | "closed", start_dt, end_dt)
    """
    now = datetime.now()
    try:
        start_dt = datetime.strptime(
            settings.get("order_start", "2099-01-01 00:00"), "%Y-%m-%d %H:%M"
        )
        end_dt = datetime.strptime(
            settings.get("order_end", "2099-01-01 00:00"), "%Y-%m-%d %H:%M"
        )
    except ValueError:
        return ("closed", None, None)

    if now < start_dt:
        return ("before", start_dt, end_dt)
    elif now > end_dt:
        return ("closed", start_dt, end_dt)
    else:
        return ("open", start_dt, end_dt)


def format_countdown(end_dt: datetime) -> str:
    """마감까지 남은 시간을 '00일 00시간 00분' 형식으로 반환합니다."""
    remaining = end_dt - datetime.now()
    if remaining.total_seconds() <= 0:
        return "마감되었습니다"
    days = remaining.days
    hours, rem = divmod(remaining.seconds, 3600)
    minutes    = rem // 60
    parts = []
    if days > 0:
        parts.append(f"{days}일")
    if hours > 0:
        parts.append(f"{hours}시간")
    parts.append(f"{minutes}분")
    return " ".join(parts)


# =============================================================================
# 입력값 검증
# =============================================================================

def validate_phone(phone: str) -> bool:
    """전화번호 형식 검증: 010-XXXX-XXXX"""
    return bool(re.match(r"^010-\d{3,4}-\d{4}$", phone.strip()))


def _fmt_phone(key: str) -> None:
    """전화번호 자동 하이픈 포맷 콜백 (on_change용).
    숫자만 입력해도 010-XXXX-XXXX 형식으로 자동 변환합니다."""
    digits = re.sub(r"\D", "", st.session_state.get(key, ""))
    if len(digits) == 11:
        st.session_state[key] = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    elif len(digits) == 10:
        st.session_state[key] = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"


# =============================================================================
# 로젠택배 엑셀 생성
# =============================================================================

def generate_logen_excel(df: pd.DataFrame, farm_name: str, settings: dict) -> bytes:
    """
    주문 DataFrame을 로젠택배 업로드 양식 엑셀로 변환합니다.

    행1 : 농장정보 텍스트 (A1 셀)
    행2 : 총 발송 집계 (품종·용도별 합계)
    행3 : 컬럼 헤더 — 수하인이름/수하인주소/수하인연락처/선물용/일반용/합계/송하인명/송하인주소/송하인연락처
    행4+: 데이터 (수신자 기준 그룹화, 주문내역 요약)
    취소 상태는 자동으로 제외됩니다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    try:
        farm_phone = st.secrets["app"]["farm_phone"]
    except Exception:
        farm_phone = settings.get("farm_phone", "")

    farm_address = settings.get("farm_address", "")

    # ── 취소·보관 상태 자동 제외 ──
    if "상태" in df.columns:
        df = df[~df["상태"].isin(["취소", "보관"])].copy()

    # ── 수량 안전 변환 (시트에서 문자열로 읽힌 경우 대비) ──
    if "수량" in df.columns:
        df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(1).astype(int)

    # ── 상품명에서 품종(약칭)·용도 분류 ──
    def _variety_abbr(s):
        s = str(s).strip()
        v = s.split()[0] if s.split() else s   # 첫 토큰 = 품종
        return v[:2]                           # 약칭 (아마쯔쿠시→아마, 경봉→경봉)
    def _usage(s):
        return "선물" if "선물용" in str(s) else ("일반" if "일반용" in str(s) else "")
    if "상품명" in df.columns:
        df["_품종"] = df["상품명"].apply(_variety_abbr)
        df["_용도"] = df["상품명"].apply(_usage)
    else:
        df["_품종"], df["_용도"] = "", ""

    # ── 용도별 총 발송 집계 (선물용/일반용/합계) ──
    total_summary = ""
    if "수량" in df.columns and "_용도" in df.columns:
        gift_tot    = int(df[df["_용도"] == "선물"]["수량"].sum())
        general_tot = int(df[df["_용도"] == "일반"]["수량"].sum())
        total_tot   = gift_tot + general_tot
        parts = []
        if gift_tot > 0:    parts.append(f"선물용-{gift_tot}박스")
        if general_tot > 0: parts.append(f"일반용-{general_tot}박스")
        if total_tot > 0:   parts.append(f"합계-{total_tot}박스")
        total_summary = ", ".join(parts)

    # ── 수신자 기준 그룹화 (대표행 + 선물용/일반용 수량 집계) ──
    group_keys = [k for k in ["받는분이름", "받는분주소", "받는분전화번호"] if k in df.columns]
    if group_keys and "수량" in df.columns:
        other_cols = [c for c in df.columns if c not in group_keys + ["수량", "_용도", "_품종"]]
        first_row  = df.groupby(group_keys, sort=False)[other_cols].first().reset_index()
        gift_qty    = (df[df["_용도"] == "선물"].groupby(group_keys, sort=False)["수량"]
                       .sum().reset_index(name="_선물용"))
        general_qty = (df[df["_용도"] == "일반"].groupby(group_keys, sort=False)["수량"]
                       .sum().reset_index(name="_일반용"))
        df = (first_row
              .merge(gift_qty,    on=group_keys, how="left")
              .merge(general_qty, on=group_keys, how="left"))
        df["_선물용"] = df["_선물용"].fillna(0).astype(int)
        df["_일반용"] = df["_일반용"].fillna(0).astype(int)
        df["_합계"]   = df["_선물용"] + df["_일반용"]

    # ── 엑셀 생성 ──
    wb = Workbook()
    ws = wb.active
    ws.title = "로젠택배"

    # 행1: 농장 정보 텍스트 (A1)
    addr_display = farm_address if farm_address else "(관리자 설정 탭에서 농장 주소를 입력해 주세요)"
    info_lines = [
        f"{farm_name}  ☎ {farm_phone}",
        f"송하인주소 : {addr_display}",
        "※ 택배비는 암호화해주세요",
    ]
    ws["A1"] = "\n".join(info_lines)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 60

    # 행2: 총 발송 집계
    ws["A2"] = f"■ 총 발송: {total_summary}" if total_summary else "■ 총 발송: (주문 없음)"
    ws["A2"].font = Font(bold=True, color="C00000")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # 행3: 컬럼 헤더
    headers = ["수하인이름", "수하인주소", "수하인연락처", "선물용", "일반용", "합계", "송하인명", "송하인주소", "송하인연락처"]
    hdr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for c_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c_idx, value=hdr)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    # 행4+: 데이터
    def _g(row_data, col):
        return row_data.get(col, "") if col in df.columns else ""

    for r_idx, (_, row_data) in enumerate(df.iterrows(), 4):
        # 본인용 주문(우리집 배달) 판별: 주문자 = 받는분(이름·전화번호 동일)
        is_self_delivery = (
            str(_g(row_data, "주문자이름")).strip() == str(_g(row_data, "받는분이름")).strip()
            and str(_g(row_data, "주문자전화번호")).strip() == str(_g(row_data, "받는분전화번호")).strip()
            and str(_g(row_data, "받는분이름")).strip() != ""
        )
        ws.cell(row=r_idx, column=1, value=_g(row_data, "받는분이름"))
        ws.cell(row=r_idx, column=2, value=_g(row_data, "받는분주소"))
        ws.cell(row=r_idx, column=3, value=_g(row_data, "받는분전화번호"))
        gift_v    = row_data.get("_선물용", 0) or 0
        general_v = row_data.get("_일반용", 0) or 0
        ws.cell(row=r_idx, column=4, value=int(gift_v)    if int(gift_v)    > 0 else "")   # 선물용
        ws.cell(row=r_idx, column=5, value=int(general_v) if int(general_v) > 0 else "")   # 일반용
        ws.cell(row=r_idx, column=6, value=int(gift_v) + int(general_v))                   # 합계
        # 선물 주문만 송하인 이름·전화번호 표기 / 본인용은 공란 / 송하인주소는 전체 공란
        ws.cell(row=r_idx, column=7, value="" if is_self_delivery else _g(row_data, "주문자이름"))      # 송하인명
        ws.cell(row=r_idx, column=8, value="")                                                          # 송하인주소 = 공란
        ws.cell(row=r_idx, column=9, value="" if is_self_delivery else _g(row_data, "주문자전화번호"))  # 송하인연락처

    # 컬럼 너비
    col_widths = [16, 36, 16, 10, 10, 10, 14, 12, 16]
    for c_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# =============================================================================
# 사이드바 — 관리자 로그인
# =============================================================================

def render_sidebar() -> bool:
    """
    사이드바에 관리자 비밀번호 입력 UI를 표시합니다.
    올바른 비밀번호 입력 시 True를 반환합니다.
    """
    # ── << 버튼 좌측 모드 라벨 (CSS 전역 주입) ──
    _is_admin = st.session_state.get("_sidebar_is_admin", False)
    _btn_lbl  = "🔧 관리자" if _is_admin else "🏠 고객"
    _btn_clr  = "#b71c1c"   if _is_admin else "#2e7d32"
    st.markdown(
        f"""<style>
        [data-testid="collapsedControl"] {{
            display: flex !important;
            flex-direction: row;
            align-items: center;
        }}
        [data-testid="collapsedControl"]::before {{
            content: "{_btn_lbl}";
            font-size: 0.72rem;
            font-weight: 700;
            color: white;
            background: {_btn_clr};
            padding: 3px 9px;
            border-radius: 10px;
            margin-right: 5px;
            white-space: nowrap;
        }}
        </style>""",
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.markdown("## 🍑 복숭아농장")
        st.markdown("---")

        # ── 비밀번호 입력 ──
        show_pw = st.checkbox("🔍 비밀번호 표시", key="show_admin_pw", value=False)
        pw = st.text_input(
            "관리자 비밀번호",
            type="default" if show_pw else "password",
            placeholder="비밀번호를 입력하세요",
            key="admin_pw",
        )

        # ── 비밀번호 미입력: 고객 모드 ──
        if not pw:
            st.markdown(
                """<div style='background:#1b5e20;border-radius:10px;
                padding:14px 16px;text-align:center;margin:10px 0;'>
                    <div style='font-size:0.72rem;color:rgba(255,255,255,0.75);
                    margin-bottom:5px;letter-spacing:1px;'>현재 모드</div>
                    <div style='font-size:1.25rem;font-weight:900;color:#ffffff;'>
                    🏠 고객 모드</div>
                    <div style='font-size:0.73rem;color:rgba(255,255,255,0.85);
                    margin-top:7px;'>관리자 전환 → 위 칸에 비밀번호 입력</div>
                </div>""",
                unsafe_allow_html=True,
            )
            st.session_state["_sidebar_is_admin"] = False
            return False

        # ── 비밀번호 검증 ──
        try:
            correct_pw = st.secrets["app"]["admin_password"]
        except Exception:
            st.error("⚠️ 관리자 비밀번호가 설정되지 않았습니다.")
            st.session_state["_sidebar_is_admin"] = False
            return False

        if pw == correct_pw:
            # 관리자 인증 완료 — force_customer 여부로 분기
            if st.session_state.get("force_customer"):
                # ── 현재: 고객 모드 (관리자 인증 완료 상태) ──
                st.markdown(
                    """<div style='background:#1b5e20;border-radius:10px;
                    padding:14px 16px;text-align:center;margin:10px 0;'>
                        <div style='font-size:0.72rem;color:rgba(255,255,255,0.75);
                        margin-bottom:5px;letter-spacing:1px;'>현재 모드</div>
                        <div style='font-size:1.25rem;font-weight:900;color:#ffffff;'>
                        🏠 고객 모드</div>
                        <div style='font-size:0.73rem;color:rgba(255,255,255,0.85);
                        margin-top:7px;'>✅ 관리자 인증 완료 · 아래 버튼으로 전환</div>
                    </div>""",
                    unsafe_allow_html=True,
                )
                if st.button("🔧 관리자 화면으로 전환 →", use_container_width=True,
                             type="primary"):
                    st.session_state["force_customer"] = False
                    st.rerun()
                st.session_state["_sidebar_is_admin"] = False
                return False
            else:
                # ── 현재: 관리자 모드 ──
                st.markdown(
                    """<div style='background:#7f0000;border-radius:10px;
                    padding:14px 16px;text-align:center;margin:10px 0;'>
                        <div style='font-size:0.72rem;color:rgba(255,255,255,0.75);
                        margin-bottom:5px;letter-spacing:1px;'>현재 모드</div>
                        <div style='font-size:1.25rem;font-weight:900;color:#ffffff;'>
                        🔧 관리자 모드</div>
                        <div style='font-size:0.73rem;color:rgba(255,255,255,0.85);
                        margin-top:7px;'>✅ 로그인 완료 · 아래 버튼으로 전환</div>
                    </div>""",
                    unsafe_allow_html=True,
                )
                if st.button("🏠 고객 화면으로 전환 →", use_container_width=True):
                    st.session_state["force_customer"] = True
                    st.rerun()
                st.session_state["_sidebar_is_admin"] = True
                return True
        else:
            # ── 비밀번호 틀림: 고객 모드 유지 ──
            st.markdown(
                """<div style='background:#1b5e20;border-radius:10px;
                padding:14px 16px;text-align:center;margin:10px 0;'>
                    <div style='font-size:0.72rem;color:rgba(255,255,255,0.75);
                    margin-bottom:5px;letter-spacing:1px;'>현재 모드</div>
                    <div style='font-size:1.25rem;font-weight:900;color:#ffffff;'>
                    🏠 고객 모드</div>
                </div>""",
                unsafe_allow_html=True,
            )
            st.error("❌ 비밀번호가 틀렸습니다")
            st.session_state["_sidebar_is_admin"] = False
            return False


# =============================================================================
# [A] 고객 주문 페이지
# =============================================================================

def _get_farm_name() -> str:
    """secrets에서 농장 이름을 읽습니다. 없으면 기본값 반환."""
    try:
        return st.secrets["app"]["farm_name"]
    except Exception:
        return "실크로드농원"


def _empty_recipient() -> dict:
    """빈 수령자 딕셔너리를 반환합니다."""
    return {"name": "", "phone": "", "address": "", "product": "", "qty": 1, "memo": ""}


def render_customer_page(settings: dict, products: list, prices: dict = None):
    """고객이 복숭아를 주문하는 공개 페이지를 렌더링합니다."""
    farm_name = _get_farm_name()

    # ── 헤더 ──
    st.markdown(
        f"<div class='peach-header'>"
        f"<h1>🍑 {farm_name}</h1>"
        f"<p>신선한 복숭아를 직접 산지에서 배달해드립니다</p>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── 주문 기간 확인 ──
    status, start_dt, end_dt = check_order_period(settings)

    if status == "before":
        st.markdown(
            f"<div class='notice-box'>"
            f"<h2>🕐 주문이 곧 시작됩니다</h2>"
            f"<p style='font-size:1.2rem;'>"
            f"주문 시작: <strong>{start_dt.strftime('%Y년 %m월 %d일 %H:%M')}</strong>"
            f"</p>"
            f"<p>잠시만 기다려 주세요! 맛있는 복숭아를 준비 중입니다. 🍑</p>"
            f"</div>",
            unsafe_allow_html=True,
        )
        return

    if status == "closed":
        end_str = end_dt.strftime("%Y년 %m월 %d일 %H:%M") if end_dt else "알 수 없음"
        st.markdown(
            f"<div class='notice-box closed'>"
            f"<h2>🚫 주문이 마감되었습니다</h2>"
            f"<p style='font-size:1.1rem;'>"
            f"마감일시: <strong>{end_str}</strong>"
            f"</p>"
            f"<p>다음 주문 기간을 기다려 주세요. 감사합니다! 🍑</p>"
            f"</div>",
            unsafe_allow_html=True,
        )

        return

    # ── 주문 접수 중: 카운트다운 표시 ──
    countdown = format_countdown(end_dt)
    st.markdown(
        f"<div class='countdown-box'>"
        f"✅ 주문 접수 중"
        f"<div style='font-size:0.88em;font-weight:normal;margin-top:4px;'>"
        f"마감까지 <strong style='font-size:1.05em;'>{countdown}</strong> 남음"
        f"</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── 주문 완료 상태 처리 ──
    if st.session_state.get("order_complete"):
        _render_order_complete(settings, farm_name)
        st.markdown("---")
        if st.button("🔄 새 주문하기"):
            st.session_state["order_complete"] = False
            st.session_state["order_result"]   = None
            st.session_state["recipients"]     = [_empty_recipient()]
            st.session_state.pop("gift_rec_ids",     None)
            st.session_state.pop("gift_rec_next_id", None)
            st.rerun()
        return

    # ── 주문 유형 선택 ──
    st.markdown("### 주문 유형을 선택해주세요")
    order_type = st.radio(
        "",
        ["🏠 우리집으로 배달 (내가 먹을려고)", "🎁 지인에게 선물"],
        key="order_type",
        horizontal=True,
        label_visibility="collapsed",
    )

    st.markdown("---")

    # ── 상품 목록 (품종별 묶음 표시, 시트 순서 유지) ──
    products_struct = load_products_struct()
    _gift_prods = products   # 선물 모드 수량 수집용

    if "우리집" in order_type:
        # ── 모드 1: 본인 수령 ──
        st.markdown("### 👤 주문자 정보")
        orderer_name  = st.text_input("이름 *",      placeholder="홍길동",              key="orderer_name")
        orderer_phone = st.text_input("전화번호 *",  placeholder="숫자만 입력 (예: 01012345678)",
                                      key="orderer_phone",
                                      on_change=_fmt_phone, args=("orderer_phone",))
        address       = st.text_input("배송 주소 *", placeholder="경북 김천시 OO로 OO", key="sender_address")
        st.markdown("### 🍑 상품 선택")
        st.caption("원하는 품종의 수량을 입력해주세요. (0박스 = 제외)")
        qtys = _render_product_qtys(products_struct, "qty_self_")
        sender_name    = orderer_name
        sender_phone   = orderer_phone
        sender_address = address
        orderer_address = address
        recipients = [
            {"name": orderer_name, "phone": orderer_phone, "address": address,
             "product": prod, "qty": qty, "memo": ""}
            for prod, qty in qtys.items() if qty > 0
        ]

    else:
        # ── 모드 2: 지인에게 선물 ──
        st.markdown("### 👤 주문자(입금자) 정보")
        st.caption("실제 입금하시는 분의 정보입니다.")
        orderer_name    = st.text_input("이름 *",     placeholder="홍길동",              key="orderer_name")
        orderer_phone   = st.text_input("전화번호 *", placeholder="숫자만 입력 (예: 01012345678)",
                                        key="orderer_phone",
                                        on_change=_fmt_phone, args=("orderer_phone",))
        orderer_address = ""   # 선물 모드: 주문자 주소 불필요

        # ── 받는 분 목록 초기화 ──
        if "gift_rec_ids" not in st.session_state:
            st.session_state["gift_rec_ids"]     = [0]
            st.session_state["gift_rec_next_id"] = 1

        rec_ids = st.session_state["gift_rec_ids"]

        st.markdown("### 🎁 받는 분 정보")
        st.caption(
            "📌 받는 분의 배송 주소가 서로 다른 경우, 각 주소마다 개별 배송됩니다. "
            "(같은 분에게 여러 상품을 보내실 때는 같은 받는 분 항목에서 수량을 올려주세요.)"
        )
        for idx, rid in enumerate(rec_ids):
            with st.container():
                if len(rec_ids) > 1:
                    col_t, col_d = st.columns([5, 1])
                    col_t.markdown(f"**📮 받는 분 {idx + 1}**")
                    if col_d.button("🗑️ 삭제", key=f"del_rec_{rid}"):
                        st.session_state["gift_rec_ids"].remove(rid)
                        for sfx in (["name", "phone", "address"]
                                    + [f"qty_{p}" for p in products]):  # 전체 products로 정리
                            st.session_state.pop(f"gr_{rid}_{sfx}", None)
                        st.rerun()
                else:
                    st.markdown("**📮 받는 분 정보**")

                st.text_input("이름 *",      placeholder="홍길동",                     key=f"gr_{rid}_name")
                st.text_input("전화번호 *",  placeholder="숫자만 입력 (예: 01012345678)",
                              key=f"gr_{rid}_phone",
                              on_change=_fmt_phone, args=(f"gr_{rid}_phone",))
                st.text_input("배송 주소 *", placeholder="서울시 강남구 테헤란로 123", key=f"gr_{rid}_address")
                st.caption("원하는 품종의 수량을 입력해주세요. (0박스 = 제외)")
                _render_product_qtys(products_struct, f"gr_{rid}_qty_")

            if idx < len(rec_ids) - 1:
                st.markdown("---")

        if st.button("➕ 받는 분 추가", use_container_width=False):
            new_id = st.session_state["gift_rec_next_id"]
            st.session_state["gift_rec_ids"].append(new_id)
            st.session_state["gift_rec_next_id"] += 1
            st.rerun()

        sender_name    = orderer_name
        sender_phone   = orderer_phone
        sender_address = orderer_address
        recipients = []
        for rid in st.session_state["gift_rec_ids"]:
            r_name    = st.session_state.get(f"gr_{rid}_name",    "")
            r_phone   = st.session_state.get(f"gr_{rid}_phone",   "")
            r_address = st.session_state.get(f"gr_{rid}_address", "")
            for prod in _gift_prods:
                qty = st.session_state.get(f"gr_{rid}_qty_{prod}", 0)
                if qty > 0:
                    recipients.append({
                        "name": r_name, "phone": r_phone,
                        "address": r_address, "product": prod,
                        "qty": qty, "memo": "",
                    })

    # ── 총 결제금액 합산 ──
    if prices and recipients:
        total_amount = sum(prices.get(r["product"], 0) * r["qty"] for r in recipients)
        if total_amount > 0:
            st.markdown(
                f"<div style='background:#fff3e0;border:2px solid #ff9800;"
                f"border-radius:12px;padding:1rem 1.5rem;margin:1rem 0;text-align:center;'>"
                f"<div style='color:#bf360c;font-size:0.9rem;margin-bottom:0.3rem;'>"
                f"💰 총 결제금액 (전체 합산)</div>"
                f"<div style='color:#e65100;font-size:2rem;font-weight:bold;'>"
                f"{total_amount:,}원</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ── 계좌 안내 ──
    st.markdown("---")
    bank   = settings.get("bank", "농협")
    acct   = settings.get("account_number", "000-0000-0000")
    holder = settings.get("holder", "장명숙")
    st.markdown(
        f"<div class='account-box'>"
        f"<p style='margin:0 0 0.3rem 0;color:#666;'>📌 주문 후 아래 계좌로 입금해 주세요</p>"
        f"<div class='bank-name'>{bank}</div>"
        f"<div class='account-num'>{acct}</div>"
        f"<div style='color:#555;'>예금주: <strong>{holder}</strong></div>"
        f"<p style='font-size:0.85rem;color:#888;margin-top:0.5rem;'>"
        f"입금자명을 주문자 성함으로 해주세요</p>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── 주문 제출 버튼 ──
    if st.button("🍑 주문 완료하기", use_container_width=True, type="primary"):
        if not recipients:
            st.error("❗ 상품을 하나 이상 선택해주세요.")
            st.stop()
        errors = _validate_order(
            orderer_name, orderer_phone,
            sender_name, sender_phone, sender_address,
            recipients,
            same_person=("우리집" in order_type),
        )
        if errors:
            for err in errors:
                st.error(err)
        else:
            _submit_order(
                orderer_name, orderer_phone,
                sender_name, sender_phone, sender_address,
                recipients, settings, farm_name,
            )


def _validate_order(orderer_name, orderer_phone, sender_name, sender_phone, sender_address, recipients, same_person=False) -> list:
    """
    주문 폼 전체 유효성 검사. 오류 메시지 리스트 반환.
    same_person=True: '우리집 배달' 모드 — 주문자=보내는분=받는분이라 중복 검증 생략.
    """
    errors = []
    # 주문자 검증 (항상)
    if not orderer_name.strip():
        errors.append("❗ 이름을 입력해주세요." if same_person else "❗ 주문자 이름을 입력해주세요.")
    if not orderer_phone.strip():
        errors.append("❗ 전화번호를 입력해주세요." if same_person else "❗ 주문자 전화번호를 입력해주세요.")
    elif not validate_phone(orderer_phone):
        errors.append("❗ 전화번호 형식이 올바르지 않습니다. (예: 010-1234-5678)")

    # 주소 검증 — 우리집 배달 모드(same_person)에서만 필수
    if same_person and not sender_address.strip():
        errors.append("❗ 배송 주소를 입력해주세요.")

    if not same_person:
        # 선물 모드: 받는 분 별도 검증 (주문자와 다른 사람)
        for i, rec in enumerate(recipients, 1):
            prefix = f"❗ {i}번째 받는 분의 " if len(recipients) > 1 else "❗ 받는 분의 "
            if not rec["name"].strip():
                errors.append(f"{prefix}이름을 입력해주세요.")
            if not rec["phone"].strip():
                errors.append(f"{prefix}전화번호를 입력해주세요.")
            elif not validate_phone(rec["phone"]):
                errors.append(f"{prefix}전화번호 형식이 올바르지 않습니다.")
            if not rec["address"].strip():
                errors.append(f"{prefix}주소를 입력해주세요.")
            if not rec.get("product"):
                errors.append(f"{prefix}상품을 선택해주세요.")
    return errors


def _submit_order(orderer_name, orderer_phone, sender_name, sender_phone, sender_address, recipients, settings, farm_name):
    """유효성 검사를 통과한 주문을 시트에 저장합니다."""
    order_number = generate_order_number()
    now_str      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 수령자별로 한 행씩 저장 (주문번호는 동일)
    rows = []
    for rec in recipients:
        rows.append([
            order_number,
            now_str,
            orderer_name.strip(),
            orderer_phone.strip(),
            sender_address.strip(),
            rec["name"].strip(),
            rec["phone"].strip(),
            rec["address"].strip(),
            rec["product"],
            str(rec["qty"]),
            rec.get("memo", "").strip(),
            "대기",
        ])

    with st.spinner("주문을 저장하는 중입니다..."):
        success = save_orders(rows)

    if success:
        load_orders.clear()   # 관리자 화면에서 즉시 새 주문을 볼 수 있도록 캐시 클리어

        # ── 주문 확인 문자 자동 발송 (CoolSMS 설정 시) ──
        product_summary = ", ".join(
            f"{r['product']} {r['qty']}박스" for r in recipients
        )
        bank   = settings.get("bank",           "농협")
        acct   = settings.get("account_number", "000-0000-0000")
        holder = settings.get("holder",         "장명숙")
        # SMS 90자 이내 유지 → SMS 요금 적용 (초과 시 LMS 자동 전환)
        sms_msg = (
            f"[복숭아농장] 주문완료\n"
            f"번호:{order_number[-6:]}\n"
            f"입금:{bank} {acct}({holder})\n"
            f"입금확인 후 발송안내드립니다."
        )
        send_sms(orderer_phone, sms_msg)   # 실패해도 주문은 완료 처리

        st.session_state["order_complete"] = True
        st.session_state["order_result"]   = {
            "order_number": order_number,
            "name":         orderer_name,
            "address":      sender_address,
            "recipients":   recipients,
        }
        st.balloons()
        st.rerun()
    else:
        st.error(
            "⚠️ 주문 저장에 실패했습니다. "
            "잠시 후 다시 시도하거나 전화로 문의해주세요."
        )


def _render_order_complete(settings: dict, farm_name: str):
    """주문 완료 화면을 렌더링합니다."""
    result       = st.session_state.get("order_result", {})
    order_number = result.get("order_number", "")
    name         = result.get("name", "")
    recipients   = result.get("recipients", [])

    st.markdown(
        f"<div class='notice-box open'>"
        f"<h2>🎉 주문이 완료되었습니다!</h2>"
        f"<p style='font-size:1.1rem;'>주문번호: <strong>{order_number}</strong></p>"
        f"<p>{name}님, 주문해주셔서 감사합니다! 🍑</p>"
        f"<p style='font-size:0.9rem;color:#555;'>"
        f"주문 접수 후 곧 메시지로 안내드리겠습니다."
        f"</p>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # 계좌 안내 재표시
    bank   = settings.get("bank", "농협")
    acct   = settings.get("account_number", "000-0000-0000")
    holder = settings.get("holder", "장명숙")
    st.markdown(
        f"<div class='account-box'>"
        f"<p style='margin:0 0 0.3rem 0;color:#666;'>💳 아래 계좌로 입금해 주세요</p>"
        f"<div class='bank-name'>{bank}</div>"
        f"<div class='account-num'>{acct}</div>"
        f"<div style='color:#555;'>예금주: <strong>{holder}</strong></div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # 주문 내역 요약 — 같은 수령자의 상품을 묶어서 표시
    st.markdown("#### 📋 주문 내역 요약")
    grouped = {}
    for rec in recipients:
        key = (rec["name"], rec["phone"], rec["address"])
        if key not in grouped:
            grouped[key] = {"name": rec["name"], "phone": rec["phone"],
                            "address": rec["address"], "memo": rec.get("memo", ""), "products": []}
        grouped[key]["products"].append(f"{rec['product']} × {rec['qty']}박스")

    for i, info in enumerate(grouped.values(), 1):
        products_html = "".join(f"<div>상품: {p}</div>" for p in info["products"])
        label = f"📮 수령자: {info['name']}" if len(grouped) == 1 else f"📮 {i}번째 수령자: {info['name']}"
        st.markdown(
            f"<div class='recipient-box'>"
            f"<div class='recipient-box-title'>{label}</div>"
            f"<div>전화번호: {info['phone']}</div>"
            f"<div>주소: {info['address']}</div>"
            f"{products_html}"
            f"</div>",
            unsafe_allow_html=True,
        )


# =============================================================================
# [B] 관리자 탭 1 — 주문 현황
# =============================================================================

def _metric_card(col, label: str, value, color: str):
    """컬럼 안에 지표 카드를 렌더링합니다."""
    with col:
        st.markdown(
            f"<div class='metric-card'>"
            f"<div class='metric-value' style='color:{color};'>{value}</div>"
            f"<div class='metric-label'>{label}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )


def render_admin_orders():
    """주문 현황 탭: 지표 요약 + 편집 가능 테이블 + CSV 다운로드"""
    st.markdown("### 📊 주문 현황")

    df = load_orders()
    if df.empty:
        st.info("ℹ️ 아직 접수된 주문이 없습니다.")
        return

    # ── 상태 값 정규화 (지표 계산 전에 먼저 실행) ──
    valid_statuses = ["대기", "입금확인", "배송준비", "배송중", "발송완료", "취소", "보관"]
    if "상태" in df.columns:
        # 알 수 없는 값 → "대기"
        df["상태"] = df["상태"].apply(lambda x: x if x in valid_statuses else "대기")

    # ── 지표 (보관 제외한 현재 시즌 기준) ──
    active_df = df[df["상태"] != "보관"] if "상태" in df.columns else df
    total    = int(active_df["주문번호"].nunique()) if "주문번호" in active_df.columns else len(active_df)
    waiting  = int((active_df["상태"] == "대기").sum())     if "상태" in active_df.columns else 0
    confirm  = int((active_df["상태"] == "입금확인").sum()) if "상태" in active_df.columns else 0
    shipped  = int((active_df["상태"].isin(["배송중", "발송완료"])).sum()) if "상태" in active_df.columns else 0
    canceled = int((active_df["상태"] == "취소").sum())     if "상태" in active_df.columns else 0
    archived = int((df["상태"] == "보관").sum())             if "상태" in df.columns else 0

    # ── 총매출 계산 (입금확인 + 배송 관련 상태 기준) ──
    revenue = 0
    prices  = load_product_prices()
    if prices and "상품명" in active_df.columns and "수량" in active_df.columns and "상태" in active_df.columns:
        paid_df = active_df[active_df["상태"].isin(["입금확인","배송준비","배송중","발송완료"])].copy()
        paid_df["수량_n"] = pd.to_numeric(paid_df["수량"], errors="coerce").fillna(0).astype(int)
        paid_df["단가"]   = paid_df["상품명"].map(prices).fillna(0).astype(int)
        revenue = int((paid_df["수량_n"] * paid_df["단가"]).sum())

    revenue_str = f"{revenue:,}원" if prices else "단가 미설정"

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    _metric_card(c1, "현재 주문",  total,        "#ff8c42")
    _metric_card(c2, "대기",       waiting,      "#ffc107")
    _metric_card(c3, "입금확인",   confirm,      "#2196f3")
    _metric_card(c4, "발송완료",   shipped,      "#4caf50")
    _metric_card(c5, "취소",       canceled,     "#9e9e9e")
    _metric_card(c6, "총매출(입금)", revenue_str, "#7b1fa2")
    _metric_card(c7, "보관(이전)",  archived,    "#78909c")

    st.markdown("---")

    # ── 검색 + 상태 필터 ──
    fc1, fc2 = st.columns([2, 1])
    with fc1:
        search_name = st.text_input("🔍 이름 검색", placeholder="주문자 또는 받는 분 이름", key="order_search")
    with fc2:
        status_view = st.selectbox(
            "상태 필터",
            options=["현재(보관제외)", "전체", "대기", "입금확인", "배송준비", "배송중", "발송완료", "취소", "보관"],
            index=0,
            key="order_status_filter",
        )

    # ── 이전 시즌 전체 보관처리 버튼 ──
    with st.expander("📦 이전 시즌 주문 보관처리"):
        st.caption("현재 시즌 주문이 끝난 후, 기존 주문 전체를 '보관' 상태로 바꿔 새 시즌 주문과 구분합니다. 보관된 데이터는 '보관' 필터로 언제든 조회할 수 있습니다.")
        if st.button("📦 현재 활성 주문 전체 보관처리", type="secondary", use_container_width=True):
            sheet = get_sheet("주문목록")
            if sheet:
                try:
                    all_rows = sheet.get_all_values()
                    if all_rows:
                        headers    = all_rows[0]
                        status_col = (headers.index("상태") + 1) if "상태" in headers else 12
                        col_idx    = {h: i for i, h in enumerate(headers)}
                        # 배치 업데이트: update_cells()로 한 번의 API 호출에 처리
                        cells_to_update = []
                        for row_i, row in enumerate(all_rows[1:], start=2):
                            cur = row[col_idx["상태"]] if "상태" in col_idx and len(row) > col_idx["상태"] else ""
                            if cur not in ("보관", "취소", ""):
                                cells_to_update.append(gspread.Cell(row_i, status_col, "보관"))
                        if cells_to_update:
                            sheet.update_cells(cells_to_update)
                            st.success(f"✅ {len(cells_to_update)}건을 보관 처리했습니다. 새 시즌 주문을 받을 준비가 됐습니다!")
                            load_orders.clear()
                            st.rerun()
                        else:
                            st.info("보관 처리할 활성 주문이 없습니다.")
                except Exception as e:
                    st.error(f"보관 처리 실패: {e}")
            else:
                st.error("Google Sheets 연결 실패")

    # ── 검색 + 상태 필터 적용 ──
    if search_name.strip():
        mask = pd.Series(False, index=df.index)
        for col in ["주문자이름", "받는분이름"]:
            if col in df.columns:
                mask |= df[col].str.contains(search_name.strip(), na=False)
        df = df[mask]
    if status_view == "현재(보관제외)" and "상태" in df.columns:
        df = df[df["상태"] != "보관"]
    elif status_view != "전체" and "상태" in df.columns:
        df = df[df["상태"] == status_view]

    # ── 최신순 정렬 ──
    if "주문일시" in df.columns:
        df = df.sort_values("주문일시", ascending=False)

    st.caption(f"표시: **{df['주문번호'].nunique() if '주문번호' in df.columns else len(df)}건** | 아래 표에서 '상태' 열을 직접 클릭하여 수정할 수 있습니다.")

    # ── 주문 순번 컬럼 추가 (같은 주문번호 = 같은 순번) ──
    if "주문번호" in df.columns:
        order_map = {v: i + 1 for i, v in enumerate(df["주문번호"].unique())}
        df.insert(0, "순번", df["주문번호"].map(order_map))
    df.index = range(1, len(df) + 1)

    # ── 자주 쓰는 '상태' 열을 맨 앞으로 이동 (가로 스크롤 없이 바로 수정 가능) ──
    front = [c for c in ["순번", "상태"] if c in df.columns]
    rest  = [c for c in df.columns if c not in front]
    df = df[front + rest]

    # ── 편집 가능한 데이터 테이블 ──
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "순번":           st.column_config.NumberColumn("순번",           disabled=True),
            "주문자이름":     st.column_config.TextColumn("보내는분이름",     disabled=True),
            "주문자전화번호": st.column_config.TextColumn("보내는분전화번호", disabled=True),
            "주문자주소":     st.column_config.TextColumn("보내는분주소",     disabled=True),
            "상태": st.column_config.SelectboxColumn(
                "상태",
                options=["대기", "입금확인", "배송준비", "배송중", "발송완료", "취소", "보관"],
            ),
        },
        key="orders_editor",
    )

    col_save, col_dl = st.columns(2)
    with col_save:
        if st.button("💾 상태 변경 저장", use_container_width=True):
            sheet = get_sheet("주문목록")
            if sheet:
                saved = 0
                try:
                    # 시트 전체를 다시 읽어 실제 행 번호를 매핑 (필터/정렬 후 오프셋 오류 방지)
                    all_rows = sheet.get_all_values()
                    if not all_rows:
                        st.warning("시트가 비어 있습니다.")
                    else:
                        headers    = all_rows[0]
                        status_col = (headers.index("상태")      + 1) if "상태"      in headers else 12
                        # (주문번호, 받는분이름, 받는분전화번호, 상품명) 조합으로 시트 행 번호 룩업 테이블 생성
                        # 전화번호 포함으로 동명이인·동일상품 다중수령 시 키 충돌 방지
                        col_idx = {h: i for i, h in enumerate(headers)}
                        lookup = {}
                        for row_i, row in enumerate(all_rows[1:], start=2):
                            def _s(c): return row[col_idx[c]] if c in col_idx and len(row) > col_idx[c] else ""
                            key = (_s("주문번호"), _s("받는분이름"), _s("받는분전화번호"), _s("상품명"))
                            lookup[key] = row_i

                        cells_to_save = []
                        for idx in range(len(df)):
                            orig_status = df.iloc[idx]["상태"] if "상태" in df.columns else ""
                            new_status  = edited_df.iloc[idx]["상태"]
                            if orig_status == new_status:
                                continue
                            row_data = edited_df.iloc[idx]
                            key = (
                                str(row_data.get("주문번호",         "")),
                                str(row_data.get("받는분이름",       "")),
                                str(row_data.get("받는분전화번호",   "")),
                                str(row_data.get("상품명",           "")),
                            )
                            if key in lookup:
                                cells_to_save.append(gspread.Cell(lookup[key], status_col, new_status))
                                saved += 1
                            else:
                                st.warning(f"시트에서 행을 찾지 못했습니다: {key}")

                        if cells_to_save:
                            sheet.update_cells(cells_to_save)
                        if saved:
                            st.success(f"✅ {saved}건의 상태를 저장했습니다.")
                            load_orders.clear()
                            st.rerun()
                        else:
                            st.info("변경된 항목이 없습니다.")
                except Exception as e:
                    st.error(f"저장 실패: {e}")
            else:
                st.error("Google Sheets 연결 실패")

    with col_dl:
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")  # BOM은 encode 시 한 번만
        st.download_button(
            "📥 CSV 다운로드",
            data=csv_bytes,
            file_name=f"주문목록_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )


# =============================================================================
# [B] 관리자 탭 2 — 주문 기간 설정
# =============================================================================

def render_admin_period(settings: dict):
    """주문 열기/닫기 탭: 즉시 열기(7일 후 자동마감) / 즉시 닫기"""
    st.markdown("### ⏰ 주문 열기 / 닫기")

    status, _, end_dt = check_order_period(settings)
    labels = {"before": "⏳ 주문 시작 전", "open": "✅ 주문 접수 중", "closed": "🚫 주문 마감"}
    st.info(f"현재 상태: **{labels.get(status, '알 수 없음')}**")
    if status == "open" and end_dt is not None:
        st.caption(f"⏳ 자동 마감 예정: {end_dt.strftime('%Y-%m-%d %H:%M')}")

    st.markdown(
        "주문을 받을 때 **[지금 바로 열기]**, 끝나면 **[지금 바로 닫기]**를 누르세요.\n\n"
        "특정 품종만 마감하려면 **[설정] 탭의 상품 관리**에서 그 품종을 '마감'으로 바꾸면 됩니다."
    )

    cb, cc = st.columns(2)
    with cb:
        if st.button("🟢 지금 바로 열기 (7일 후 자동마감)", use_container_width=True, type="primary"):
            now = datetime.now()
            settings["order_start"] = now.strftime("%Y-%m-%d %H:%M")
            settings["order_end"]   = (now + timedelta(days=7)).strftime("%Y-%m-%d %H:%M")
            if save_settings(settings):
                load_settings.clear()
                st.success("✅ 주문이 열렸습니다! (7일 후 자동 마감)")
                st.rerun()
            else:
                st.error("저장 실패")

    with cc:
        if st.button("🔴 지금 바로 닫기", use_container_width=True):
            past = datetime.now() - timedelta(minutes=1)
            settings["order_end"] = past.strftime("%Y-%m-%d %H:%M")
            if save_settings(settings):
                load_settings.clear()
                st.success("✅ 주문이 마감되었습니다.")
                st.rerun()
            else:
                st.error("저장 실패")


# =============================================================================
# [B] 관리자 탭 3 — 로젠택배 엑셀
# =============================================================================

def render_admin_logen(settings: dict):
    """로젠택배 엑셀 생성 탭: 필터 → 미리보기 → 다운로드 → 발송완료 처리"""
    st.markdown("### 📦 로젠택배 발송 엑셀")

    df = load_orders()
    if df.empty:
        st.info("ℹ️ 주문 데이터가 없습니다.")
        return

    # ── 필터 ──
    col1, col2 = st.columns(2)
    with col1:
        status_filter = st.multiselect(
            "상태 필터",
            options=["대기", "입금확인", "발송완료", "취소"],
            default=["대기", "입금확인"],
            key="logen_status",
        )
    with col2:
        date_range = None
        if "주문일시" in df.columns:
            try:
                df["_dt"] = pd.to_datetime(df["주문일시"], errors="coerce")
                min_d = df["_dt"].min().date()
                max_d = df["_dt"].max().date()
                date_range = st.date_input("주문일 범위", value=(min_d, max_d), key="logen_date")
            except Exception:
                pass

    # 필터 적용
    filtered = df.copy()
    if "상태" in df.columns and status_filter:
        filtered = filtered[filtered["상태"].isin(status_filter)]
    if date_range and len(date_range) == 2 and "_dt" in filtered.columns:
        s_d, e_d = date_range
        filtered = filtered[
            (filtered["_dt"].dt.date >= s_d) &
            (filtered["_dt"].dt.date <= e_d)
        ]

    st.caption(f"필터 결과: **{len(filtered)}건**")
    if filtered.empty:
        st.warning("조건에 맞는 주문이 없습니다.")
        return

    # 미리보기
    # 택배 준비에 필요한 칸만 표시 (주문번호·상태는 위 필터로 이미 구분되고
    # 택배회사 엑셀에도 포함되지 않으므로 미리보기에서 제외)
    preview_cols = [c for c in [
        "주문자이름", "주문자전화번호", "주문자주소",
        "받는분이름", "받는분전화번호", "받는분주소",
        "상품명", "수량",
    ] if c in filtered.columns]
    preview_df = filtered[preview_cols].rename(columns={
        "주문자이름":     "보내는분이름",
        "주문자전화번호": "보내는분전화번호",
        "주문자주소":     "보내는분주소",
    })
    st.dataframe(preview_df, use_container_width=True)

    farm_name = _get_farm_name()
    col_dl, col_mark = st.columns(2)

    with col_dl:
        excel_bytes = generate_logen_excel(filtered, farm_name, settings)
        st.download_button(
            "📥 로젠택배 엑셀 다운로드",
            data=excel_bytes,
            file_name=f"로젠택배_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_mark:
        st.caption("엑셀로 다운받아 택배회사에 접수완료한 건에 대하여 일괄 발송완료 처리합니다.")
        # 대상 건수(이미 발송완료/취소 제외) 미리 계산
        if "상태" in filtered.columns:
            eligible_n = int((~filtered["상태"].isin(["발송완료", "취소"])).sum())
        else:
            eligible_n = len(filtered)

        # 1단계 안전장치: 확인 체크박스를 켜야만 버튼 활성화
        confirm_chk = st.checkbox("위 내용을 확인했습니다", key="logen_done_confirm_chk")

        clicked = st.button(
            "✅ 선택 주문 '발송완료' 처리",
            use_container_width=True,
            disabled=not confirm_chk,
        )

        if clicked and not st.session_state.get("logen_done_pending"):
            # 2단계 안전장치: 첫 클릭은 확인 요청만 (확정 안 함)
            st.session_state["logen_done_pending"] = True
            st.warning(
                f"⚠️ {eligible_n}건을 '발송완료' 처리합니다. "
                "확정하려면 버튼을 한 번 더 눌러주세요."
            )
        elif clicked and st.session_state.get("logen_done_pending"):
            # 두 번째 클릭 → 실제 처리 실행
            st.session_state["logen_done_pending"] = False
            sheet = get_sheet("주문목록")
            if sheet:
                try:
                    all_rows = sheet.get_all_values()
                    if not all_rows:
                        st.warning("시트가 비어있습니다.")
                    else:
                        # 헤더에서 열 번호를 동적으로 계산
                        headers    = all_rows[0]
                        col_idx    = {h: i for i, h in enumerate(headers)}
                        status_col = (headers.index("상태") + 1) if "상태" in headers else 12
                        status_idx = status_col - 1
                        order_idx  = col_idx.get("주문번호", 0)  # 하드코딩 제거
                        target_ids = set(filtered["주문번호"].tolist())
                        cells_done = []
                        for row_idx, row in enumerate(all_rows[1:], start=2):
                            if (len(row) > max(status_idx, order_idx)
                                    and row[order_idx] in target_ids
                                    and row[status_idx] not in ("발송완료", "취소")):
                                cells_done.append(gspread.Cell(row_idx, status_col, "발송완료"))
                        if cells_done:
                            sheet.update_cells(cells_done)
                            st.success(f"✅ {len(cells_done)}건을 '발송완료' 처리했습니다.")
                            load_orders.clear()
                            st.rerun()
                        else:
                            st.info("변경할 항목이 없습니다. (이미 발송완료/취소 상태)")
                except Exception as e:
                    st.error(f"처리 실패: {e}")
            else:
                st.error("Google Sheets 연결 실패")
        elif st.session_state.get("logen_done_pending"):
            # 확인 대기 상태 표시 (버튼을 누르지 않은 다른 재실행 시)
            st.warning(
                f"⚠️ {eligible_n}건 발송완료 대기 중 — 버튼을 한 번 더 누르면 확정됩니다."
            )


# =============================================================================
# [B] 관리자 탭 4 — 메시지 발송 안내
# =============================================================================

def render_admin_email(settings: dict):
    """고객 공지 메시지 발송 탭 — 문자/카카오톡 발송용 문구 생성"""
    st.markdown("### 💬 고객 공지 메시지 발송")

    farm_name = _get_farm_name()
    try:
        order_url = st.secrets["app"]["order_url"]
    except Exception:
        order_url = "https://your-app.streamlit.app"

    order_start = settings.get("order_start", "2026-07-01 09:00")
    order_end   = settings.get("order_end",   "2026-07-10 18:00")
    bank        = settings.get("bank",         "농협")
    acct        = settings.get("account_number", "000-0000-0000")
    holder      = settings.get("holder",         "장명숙")

    st.info("아래 메시지를 복사하여 카카오톡, 문자 등으로 고객에게 보내주세요.")

    # ── 메시지 템플릿 ──
    st.markdown("**메시지 내용 작성**")
    default_body = (
        f"안녕하세요! 🍑\n"
        f"{farm_name}입니다.\n\n"
        f"올해도 맛있는 복숭아 주문을 받습니다.\n\n"
        f"▶ 주문 기간: {order_start} ~ {order_end}\n"
        f"▶ 주문 링크: {order_url}\n\n"
        f"입금 계좌 안내\n"
        f"─────────────\n"
        f"은행: {bank}\n"
        f"계좌: {acct}\n"
        f"예금주: {holder}\n"
        f"─────────────\n\n"
        f"신선하고 달콤한 복숭아로\n"
        f"무더운 여름 보내세요! 🍑\n\n"
        f"감사합니다.\n"
        f"{farm_name}"
    )

    # 저장된 메시지가 있으면 기본값으로 사용, 없으면 자동 생성 문구 사용
    saved_body = settings.get("notice_message", "")
    init_body  = saved_body if saved_body.strip() else default_body
    body = st.text_area("메시지 본문 (수정 가능)", value=init_body, height=320, key="msg_body")

    # ── 메시지 내용 저장 / 초기화 ──
    col_save, col_reset = st.columns(2)
    with col_save:
        if st.button("💾 메시지 내용 저장", use_container_width=True):
            if body.strip():
                settings["notice_message"] = body
                if save_settings(settings):
                    load_settings.clear()
                    st.success("메시지 내용이 저장되었습니다. 다음에 열어도 이 내용이 유지됩니다.")
                else:
                    st.error("저장 실패 - Google Sheets 연결을 확인해주세요.")
            else:
                st.warning("저장할 메시지 내용을 입력해주세요.")
    with col_reset:
        if st.button("↩️ 기본 문구로 초기화", use_container_width=True):
            settings.pop("notice_message", None)
            save_settings(settings)
            load_settings.clear()
            st.session_state.pop("msg_body", None)
            st.rerun()
    if saved_body.strip():
        st.caption("💾 저장된 메시지를 불러왔습니다.")

    st.markdown("---")

    # ── 고객 목록 및 문자 발송 ──
    customers_df = load_customers()
    if not customers_df.empty:
        st.markdown(f"**📋 고객 목록 ({len(customers_df)}명)**")
        st.dataframe(customers_df, use_container_width=True)

        st.markdown("---")
        # CoolSMS 설정 여부 확인
        sms_ready = _COOLSMS_AVAILABLE and ("coolsms" in st.secrets)
        if sms_ready:
            st.markdown("**📱 문자 일괄 발송 (CoolSMS)**")
            # 전화번호 컬럼 자동 감지
            phone_col = next(
                (c for c in customers_df.columns if "전화" in c or "phone" in c.lower()),
                None
            )
            if phone_col:
                numbers = customers_df[phone_col].dropna().tolist()
                st.caption(f"발송 대상: **{len(numbers)}명** · 발송 전 메시지를 반드시 확인하세요.")
                col_send, col_test = st.columns(2)
                with col_send:
                    if st.button("📤 전체 고객에게 문자 발송", use_container_width=True,
                                 type="primary"):
                        if not body.strip():
                            st.error("메시지 내용을 입력해주세요.")
                        else:
                            with st.spinner(f"{len(numbers)}명에게 발송 중..."):
                                ok, fail = send_sms_bulk(numbers, body)
                            st.success(f"발송 완료: 성공 {ok}건 / 실패 {fail}건")
                with col_test:
                    test_number = st.text_input("테스트 번호", placeholder="01012345678",
                                                key="sms_test_num")
                    if st.button("테스트 문자 발송", use_container_width=True):
                        if test_number.strip():
                            if send_sms(test_number.strip(), body):
                                st.success(f"테스트 문자 발송 완료 ({test_number})")
                            else:
                                st.error("발송 실패 - CoolSMS 설정을 확인해주세요.")
                        else:
                            st.warning("테스트 번호를 입력해주세요.")
            else:
                st.warning("고객목록 시트에 '전화번호' 열이 없습니다.")
        else:
            st.markdown("**직접 발송 안내**")
            if not _COOLSMS_AVAILABLE:
                st.info(
                    "자동 문자 발송을 사용하려면:\n"
                    "1. `pip install coolsms` 설치\n"
                    "2. secrets.toml에 [coolsms] api_key / api_secret / sender 추가\n\n"
                    "지금은 위 메시지를 복사해서 카카오톡/문자로 직접 보내주세요."
                )
            else:
                st.info("secrets.toml에 [coolsms] 설정을 추가하면 자동 발송이 활성화됩니다.")
    else:
        st.warning("고객목록 시트에 고객 정보가 없습니다. '이름 / 전화번호' 형태로 입력해주세요.")


# =============================================================================
# [B] 관리자 탭 5 - 설정
# =============================================================================

def render_admin_settings(settings: dict):
    """농장 정보 및 앱 설정 탭"""
    st.markdown("### 설정")

    st.markdown("**농장 / 계좌 정보**")
    col1, col2 = st.columns(2)
    with col1:
        new_bank    = st.text_input("은행명",        value=settings.get("bank",           "농협"),   key="cfg_bank")
        new_holder  = st.text_input("예금주",        value=settings.get("holder",         "장명숙"), key="cfg_holder")
    with col2:
        new_acct    = st.text_input("계좌번호",      value=settings.get("account_number", ""),      key="cfg_acct")
        new_phone   = st.text_input("농장 전화번호", value=settings.get("farm_phone",     ""),      key="cfg_phone")
    new_address = st.text_input(
        "농장 주소 (로젠택배 송하인 주소)",
        value=settings.get("farm_address", ""),
        placeholder="경북 김천시 감문면 문무리 1269번지",
        key="cfg_farm_address",
    )

    if st.button("설정 저장"):
        settings.update({
            "bank":           new_bank,
            "holder":         new_holder,
            "account_number": new_acct,
            "farm_phone":     new_phone,
            "farm_address":   new_address,
        })
        if save_settings(settings):
            load_settings.clear()
            st.success("설정이 저장되었습니다.")
            st.rerun()
        else:
            st.error("저장 실패. Google Sheets 연결을 확인해주세요.")

    # ── 상품 / 단가 관리 ──
    st.markdown("---")
    st.markdown("**상품 / 단가 관리**")
    st.caption(
        "상품명은 '품종 4kg 선물용/일반용' 형식으로 입력하세요 (예: 아마쯔쿠시 4kg 선물용). "
        "맨 아래 빈 줄에 입력하면 상품이 추가되고, 줄 왼쪽을 선택해 삭제할 수 있습니다. "
        "특정 품종을 그만 받으려면 '상태'를 마감으로 바꾸세요. "
        "수정 후 반드시 아래 [상품·단가 저장] 버튼을 눌러주세요."
    )
    prod_df = load_products_full()
    if not prod_df.empty:
        prod_df["단가"] = (
            pd.to_numeric(
                prod_df["단가"].astype(str).str.replace(",", "", regex=False)
                                          .str.replace("원", "", regex=False).str.strip(),
                errors="coerce",
            ).fillna(0).astype(int)
        )
    edited_prod = st.data_editor(
        prod_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "상품명": st.column_config.TextColumn("상품명", required=True),
            "단가":   st.column_config.NumberColumn("단가(원)", min_value=0, step=1000, format="%d"),
            "설명":   st.column_config.TextColumn("설명"),
            "상태":   st.column_config.SelectboxColumn("상태", options=["판매중", "마감", "예정"], default="판매중", required=True),
        },
        key="products_editor",
    )
    if st.button("상품·단가 저장", key="save_products_btn"):
        if save_products(edited_prod):
            load_products.clear()
            load_product_prices.clear()
            load_products_full.clear()
            load_products_struct.clear()
            st.success("상품 정보가 저장되었습니다.")
            st.rerun()
        else:
            st.error("저장 실패. Google Sheets 연결을 확인해주세요.")

    st.markdown("---")
    st.markdown("**secrets.toml 설정 안내**")
    st.info(
        "아래 항목은 `.streamlit/secrets.toml` 파일 (또는 Streamlit Cloud Secrets)에서 설정합니다.\n\n"
        "- `[gcp_service_account]` - Google 서비스 계정 JSON 내용\n"
        "- `[app] spreadsheet_id` - Google Sheets URL의 `/d/` 뒤 ID\n"
        "- `[app] admin_password` - 관리자 비밀번호\n"
        "- `[coolsms]` - CoolSMS API 키 (문자 자동발송용)\n"
        "- `[account]` - 기본 계좌 정보"
    )

    st.markdown("---")
    st.markdown("**Google Sheets 연결 상태**")
    client = get_gspread_client()
    if client is not None:
        ss = get_spreadsheet()
        if ss is not None:
            st.success("Google Sheets 연결 성공")
            try:
                ws_names = [ws.title for ws in ss.worksheets()]
                st.write("시트 목록:", ", ".join(ws_names))
                required = {"주문목록", "상품목록", "고객목록", "설정"}
                missing  = required - set(ws_names)
                if missing:
                    st.warning(f"누락된 시트: {', '.join(missing)}")
                else:
                    st.success("필수 시트(주문목록/상품목록/고객목록/설정) 모두 존재")
            except Exception:
                pass
        else:
            st.error("스프레드시트를 열 수 없습니다. spreadsheet_id를 확인해주세요.")
    else:
        st.error("Google Sheets 연결 실패. gcp_service_account 설정을 확인해주세요.")

    # ── ⚠️ 주문 데이터 초기화 (위험 구역) ──
    st.markdown("---")
    st.markdown("**⚠️ 주문 데이터 초기화**")
    st.caption(
        "테스트 데이터를 지우거나 새 시즌을 시작할 때 사용하세요. "
        "주문목록의 모든 주문이 삭제되며 되돌릴 수 없습니다. "
        "삭제 전 반드시 아래에서 백업(CSV)을 먼저 받아두세요. (상품·계좌·설정은 삭제되지 않습니다.)"
    )
    backup_df = load_orders()
    if not backup_df.empty:
        st.caption(f"현재 저장된 주문: {len(backup_df)}줄")
        st.download_button(
            "📥 삭제 전 백업 다운로드 (CSV)",
            data=backup_df.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"주문백업_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            key="reset_backup_dl",
        )
    else:
        st.caption("현재 저장된 주문이 없습니다.")

    reset_chk = st.checkbox(
        "백업을 받았으며, 모든 주문을 삭제하는 데 동의합니다", key="reset_orders_chk"
    )
    reset_clicked = st.button(
        "🗑️ 주문 데이터 전체 삭제", disabled=not reset_chk, key="reset_orders_btn"
    )
    if reset_clicked and not st.session_state.get("reset_orders_pending"):
        st.session_state["reset_orders_pending"] = True
        st.warning("⚠️ 모든 주문을 삭제합니다. 확정하려면 버튼을 한 번 더 눌러주세요.")
    elif reset_clicked and st.session_state.get("reset_orders_pending"):
        st.session_state["reset_orders_pending"] = False
        sheet = get_sheet("주문목록")
        if sheet:
            try:
                all_rows = sheet.get_all_values()
                header = all_rows[0] if all_rows else [
                    "주문번호", "주문일시", "주문자이름", "주문자전화번호", "주문자주소",
                    "받는분이름", "받는분전화번호", "받는분주소",
                    "상품명", "수량", "배송메모", "상태",
                ]
                sheet.clear()
                sheet.update("A1", [header])
                load_orders.clear()
                st.success("모든 주문 데이터를 삭제했습니다. (제목 줄은 유지)")
                st.rerun()
            except Exception as e:
                st.error(f"삭제 실패: {e}")
        else:
            st.error("Google Sheets 연결 실패")
    elif st.session_state.get("reset_orders_pending"):
        st.warning("⚠️ 삭제 대기 중 — 버튼을 한 번 더 누르면 모든 주문이 삭제됩니다.")


# =============================================================================
# [B] 관리자 페이지 - 탭 통합
# =============================================================================

def render_admin_page(settings: dict, products: list):
    """관리자 모드 전체 레이아웃: 헤더 + 5개 탭"""
    farm_name = _get_farm_name()
    st.markdown(
        f"<div class='peach-header'>"
        f"<h1>{farm_name}<span class='admin-sub'>관리자 페이지</span></h1>"
        f"<p>주문 관리 및 운영 설정 페이지</p>"
         f"</div>",
        unsafe_allow_html=True,
    )

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "주문 현황",
        "주문 기간 설정",
        "로젠택배 엑셀",
        "메시지 발송",
        "설정",
    ])

    with tab1: render_admin_orders()
    with tab2: render_admin_period(settings)
    with tab3: render_admin_logen(settings)
    with tab4: render_admin_email(settings)
    with tab5: render_admin_settings(settings)


# =============================================================================
# 메인 진입점
# =============================================================================

def main():
    """앱 진입점: 세션 초기화 -> 사이드바 로그인 -> 관리자/고객 분기"""

    # 모바일 스크롤 + 핀치줌 허용
    import streamlit.components.v1 as _components
    _components.html(
        """
        <script>
        (function() {
            try {
                var doc = window.parent.document;
                // 뷰포트: 핀치줌 허용
                var meta = doc.querySelector('meta[name="viewport"]');
                if (meta) {
                    meta.setAttribute('content',
                        'width=device-width, initial-scale=1.0, maximum-scale=5.0, user-scalable=yes');
                }
                // 스크롤 허용 (touch 이벤트가 스크롤을 막지 않도록)
                doc.addEventListener('touchmove', function(e) {
                    e.stopPropagation();
                }, { passive: true, capture: true });
                // 스타일: 모든 요소에 핀치줌 + 스크롤 허용 (중복 방지)
                if (!doc.getElementById('peach-touch-fix')) {
                    var style = doc.createElement('style');
                    style.id = 'peach-touch-fix';
                    style.textContent =
                        '* { touch-action: pan-y pinch-zoom !important; }' +
                        'html, body { overflow-y: auto !important; -webkit-overflow-scrolling: touch !important; }';
                    doc.head.appendChild(style);
                }
            } catch(e) {}
        })();
        </script>
        """,
        height=0,
    )

    for key, default in [
        ("order_complete", False),
        ("order_result",   None),
        ("recipients",     None),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    is_admin = render_sidebar()

    settings = load_settings()
    products = load_products()
    prices   = load_product_prices()

    if is_admin:
        render_admin_page(settings, products)
    else:
        render_customer_page(settings, products, prices)


if __name__ == "__main__":
    main()
